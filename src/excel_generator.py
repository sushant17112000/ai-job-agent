"""
Excel Generator — produces a formatted, color-coded job report workbook.
"""

import logging
from collections import Counter
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Color palette
COLOR_HEADER_FILL = "1F3864"   # dark blue
COLOR_HEADER_FONT = "FFFFFF"   # white
COLOR_GREEN = "C6EFCE"         # score >= 80
COLOR_YELLOW = "FFEB9C"        # score 60-79
COLOR_GREEN_FONT = "276221"
COLOR_YELLOW_FONT = "9C6500"


def _auto_width(ws):
    """Set column widths based on maximum cell content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        # Cap width at 60 characters, minimum 12
        adjusted = min(max(max_len + 2, 12), 60)
        ws.column_dimensions[col_letter].width = adjusted


def generate_excel(
    scored_jobs: list[dict],
    cv_profile: dict,
    report_date: date | None = None,
    output_dir: str = "reports",
) -> str:
    """
    Build and save a formatted Excel workbook.

    Args:
        scored_jobs: list of job dicts with match_score, match_reason, rank fields.
        cv_profile: parsed CV profile dict.
        report_date: date for the filename (defaults to today).
        output_dir: directory to save the file.

    Returns:
        Absolute path to the saved xlsx file.
    """
    if report_date is None:
        report_date = date.today()

    Path(output_dir).mkdir(parents=True, exist_ok=True)
    filename = f"{report_date.isoformat()}_jobs.xlsx"
    filepath = Path(output_dir) / filename

    # --- Build main DataFrame ---
    columns = [
        "Rank",
        "Job Title",
        "Company",
        "Location",
        "Portal",
        "Match Score",
        "Match Reason",
        "Apply URL",
        "Date Posted",
    ]

    rows = []
    for job in scored_jobs:
        rows.append(
            {
                "Rank": job.get("rank", ""),
                "Job Title": job.get("title", ""),
                "Company": job.get("company", ""),
                "Location": job.get("location", ""),
                "Portal": job.get("portal", ""),
                "Match Score": job.get("match_score", ""),
                "Match Reason": job.get("match_reason", ""),
                "Apply URL": job.get("url", ""),
                "Date Posted": job.get("date_posted", ""),
            }
        )

    df_jobs = pd.DataFrame(rows, columns=columns) if rows else pd.DataFrame(columns=columns)
    df_jobs.to_excel(str(filepath), index=False, sheet_name="Jobs")

    # --- Apply formatting with openpyxl ---
    wb = load_workbook(str(filepath))
    ws_jobs = wb["Jobs"]

    header_fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_FILL)
    header_font = Font(bold=True, color=COLOR_HEADER_FONT)

    # Style header row
    for cell in ws_jobs[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_jobs.freeze_panes = "A2"

    # Find column indices
    col_map = {cell.value: cell.column for cell in ws_jobs[1]}
    score_col = col_map.get("Match Score")
    url_col = col_map.get("Apply URL")

    # Style data rows
    for row in ws_jobs.iter_rows(min_row=2):
        score_cell = row[score_col - 1] if score_col else None
        score = score_cell.value if score_cell else 0

        if isinstance(score, (int, float)) and score >= 80:
            fill = PatternFill(fill_type="solid", fgColor=COLOR_GREEN)
            font_color = COLOR_GREEN_FONT
        elif isinstance(score, (int, float)) and score >= 60:
            fill = PatternFill(fill_type="solid", fgColor=COLOR_YELLOW)
            font_color = COLOR_YELLOW_FONT
        else:
            fill = None
            font_color = "000000"

        for cell in row:
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.column != url_col:
                cell.font = Font(color=font_color)

        # Make URL a clickable hyperlink
        if url_col:
            url_cell = row[url_col - 1]
            url_val = url_cell.value
            if url_val:
                url_cell.hyperlink = url_val
                url_cell.font = Font(color="0563C1", underline="single")
                url_cell.value = "Apply Here"

    _auto_width(ws_jobs)

    # --- Summary Sheet ---
    portal_counts = Counter(job.get("portal", "Unknown") for job in scored_jobs)
    location_counts = Counter(job.get("location", "Unknown") for job in scored_jobs)
    top_companies = Counter(job.get("company", "Unknown") for job in scored_jobs).most_common(5)

    ws_summary = wb.create_sheet("Summary")

    summary_data = [
        ["AI Job Search Report — Summary"],
        [],
        ["Report Date", report_date.isoformat()],
        ["Candidate", cv_profile.get("name", "Unknown")],
        ["Total Matched Jobs", len(scored_jobs)],
        [],
        ["Jobs by Portal"],
    ]
    for portal, count in portal_counts.items():
        summary_data.append([f"  {portal}", count])
    summary_data.append([])
    summary_data.append(["Jobs by Location (Top 10)"])
    for loc, count in location_counts.most_common(10):
        summary_data.append([f"  {loc}", count])
    summary_data.append([])
    summary_data.append(["Top 5 Companies"])
    for company, count in top_companies:
        summary_data.append([f"  {company}", count])

    for row_data in summary_data:
        ws_summary.append(row_data)

    # Style the title
    ws_summary["A1"].font = Font(bold=True, size=14, color=COLOR_HEADER_FILL)
    _auto_width(ws_summary)

    wb.save(str(filepath))
    logger.info("Excel report saved: %s (%d jobs)", filepath, len(scored_jobs))
    return str(filepath.resolve())
